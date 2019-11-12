from bs4 import BeautifulSoup
from urllib.request import urlopen

from datetime import date
from openpyxl import load_workbook

from Books import nationalBook, internationalBook

# Global vars
urlToParse = u"https://www.amazon.com.br/Resist%C3%AAncia-dos-materiais-R-Hibbeler-ebook/dp/B07WH4677K/ref=sr_1_1?qid=1573499147&refinements=p_27%3AR.+C.+Hibbeler&s=digital-text&sr=1-1"
urlToParse = u"https://www.amazon.com.br"

originalSpeadsheet = "aquisicaodelivros.xlsx"

DEBUG_FLAG = False

class Request:
    def __init__(self):
        self.library    = None
        self.branchLine = None
        self.email      = None
        self.email      = None
        self.date       = None
        self.books      = []

    def getMetaData(self):
        self.library        = input("Insira o nome da Biblioteca: ")
        self.branchLine     = input("Insira o ramal: ")
        self.email          = input("Insira o seu e-mail: ")
        self.responsible    = input("Insira o nome do responsável pela solicitação: ")
        self.date           = date.today()

    def getLinks(self):
        #TODO implement
        pass

    def appendBook(self, book):
        self.books.append(book)

    def __str__(self):
        out = ""
        out += "This is a resquest. Attrs:"
        out += str(vars(self))
        return out

def request2workbook(r, wb):
    nationalBooks       = [n for n in r.books if not n.internationalFlag]
    internationalBooks  = [n for n in r.books if n.internationalFlag]


    def fillRow(row, book):
        #for i in columns:
        #ws[i][row] = book.
        pass


    # Fill national sheet
    wb.active = (wb.sheetnames.index('Nacionais'));     ws = wb.active

    emptyCell =  [cell.value for cell in ws['A']].index('UNIDADE')
    emptyCell += [cell.value for cell in ws['A'][emptyCell:]].index(None)

    for i in range(len(nationalBooks)):
        fillRow(emptyCell+i, nationalBooks[i])

    # Fill international sheet
    wb.active = (wb.sheetnames.index('Estrangeiros'));  ws = wb.active

    emptyCell =  [cell.value for cell in ws['A']].index('UNIDADE')
    emptyCell += [cell.value for cell in ws['A'][emptyCell:]].index(None)

    for i in range(len(internationalBooks)):
        fillRow(emptyCell+i, internationalBooks[i])


def openUrlAndSoup(url):
    #TODO catch para urllib.error.URLError
    with urlopen(url) as response:
        soup = BeautifulSoup(response, 'html.parser')
    return soup

def anchorAndGet(soup):
    # TODO entender isso aqui:
    for anchor in soup.find_all('a'):
        a = anchor.get('href', '/')
    print(type(a))

def debug(str):
    if DEBUG_FLAG:
        print(str)

def main():
    request = Request()
    #request.getMetaData()
    #print(request)

    wb = load_workbook(originalSpeadsheet)
    print(wb.sheetnames)

    brBook = nationalBook(unidade='FEM', seq=1, author='Machado', title='Dom Casmurro', publisher="Unicamp", vol=1,
                          year=1899, edition=1, ISBN10="?"*10, ISBN13="?"*13, priority=1, quantity=1, unitValue=20,
                          observation="LivroBR")

    brBook2 = nationalBook(unidade='FEM', seq=1, author='Machado', title='Memórias Póstumas', publisher="Unicamp", vol=1,
                          year=1881, edition=1, ISBN10="?"*10, ISBN13="?"*13, priority=1, quantity=1, unitValue=20,
                          observation="LivroBR")

    euaBook = internationalBook(unidade='FEM', seq=1, author='J. K. Rowling', title='Harry Potter', publisher="Rocco",
                                vol=1, year=1997, edition=1, ISBN10="!"*10, ISBN13="!"*13, priority=2, quantity=1,
                                unitValue=20, dollarFlag=True, observation="Livro Gringo")

    print(str(brBook))
    print(str(euaBook))
    print(str(brBook2))

    print(request)
    #TODO getUrls
    #TODO for url in urls:
    #    soup = openUrlAndSoup(urlToParse)
    #    anchorAndGet(soup) TODO: return useful data

    request.appendBook(brBook)
    request.appendBook(euaBook)
    request.appendBook(brBook2)

    request2workbook(request, wb)

    #TODO save workbook

if __name__ == "__main__":
    main()